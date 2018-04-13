#!/usr/bin/env python
#coding:utf-8
###  pip install XlsxWriter==1.0.2 lxml==4.1.1 Wand==0.4.4 openpyxl==2.4.8

import zipfile
import glob
import os
import re
import time
import shutil
import zipfile
from lxml import etree
from collections import OrderedDict
from wand.image import Image
from os.path import abspath, basename, join
import urllib,argparse
from openpyxl import load_workbook
from copy import copy

def opendocx(file):
    '''Open a docx file, return a document XML tree'''
    mydoc = zipfile.ZipFile(file)
    xmlcontent = mydoc.read('word/document.xml')
    document = etree.fromstring(xmlcontent)
    return document
    
nsprefixes = {'a': 'http://schemas.openxmlformats.org/drawingml/2006/main',
 'cp': 'http://schemas.openxmlformats.org/package/2006/metadata/core-properties',
 'ct': 'http://schemas.openxmlformats.org/package/2006/content-types',
 'cx': 'http://schemas.microsoft.com/office/drawing/2014/chartex',
 'dc': 'http://purl.org/dc/elements/1.1/',
 'dcmitype': 'http://purl.org/dc/dcmitype/',
 'dcterms': 'http://purl.org/dc/terms/',
 'ep': 'http://schemas.openxmlformats.org/officeDocument/2006/extended-properties',
 'm': 'http://schemas.openxmlformats.org/officeDocument/2006/math',
 'mc': 'http://schemas.openxmlformats.org/markup-compatibility/2006',
 'mo': 'http://schemas.microsoft.com/office/mac/office/2008/main',
 'mv': 'urn:schemas-microsoft-com:mac:vml',
 'o': 'urn:schemas-microsoft-com:office:office',
 'pic': 'http://schemas.openxmlformats.org/drawingml/2006/picture',
 'pr': 'http://schemas.openxmlformats.org/package/2006/relationships',
 'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
 'v': 'urn:schemas-microsoft-com:vml',
 've': 'http://schemas.openxmlformats.org/markup-compatibility/2006',
 'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main',
 'w10': 'urn:schemas-microsoft-com:office:word',
 'w14': 'http://schemas.microsoft.com/office/word/2010/wordml',
 'w15': 'http://schemas.microsoft.com/office/word/2012/wordml',
 'w16se': 'http://schemas.microsoft.com/office/word/2015/wordml/symex',
 'wne': 'http://schemas.microsoft.com/office/word/2006/wordml',
 'wp': 'http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing',
 'wp14': 'http://schemas.microsoft.com/office/word/2010/wordprocessingDrawing',
 'wpc': 'http://schemas.microsoft.com/office/word/2010/wordprocessingCanvas',
 'wpg': 'http://schemas.microsoft.com/office/word/2010/wordprocessingGroup',
 'wpi': 'http://schemas.microsoft.com/office/word/2010/wordprocessingInk',
 'wps': 'http://schemas.microsoft.com/office/word/2010/wordprocessingShape',
 'xsi': 'http://www.w3.org/2001/XMLSchema-instance'}
 
docnsmap = {i:nsprefixes[i] for i in ['ve', 'wne', 'v', 'm', 'o', 'w10', 'r', 'w', 'wp']}
corensmap = {i:nsprefixes[i] for i in ['dcterms', 'dcmitype', 'cp', 'dc', 'xsi']}
    
def makeelement(tagname, tagtext=None, nsprefix='w', attributes=None, attrnsprefix="w", namespace=None):
    '''makeelement("r",tagtext="ABC",nsprefix="w",attributes={"rsidR":"00C14938"},attrnsprefix="w",namespace=root.nsmap)'''  
    full_attr = {}
    docnsmap = {i:nsprefixes[i] for i in ['ve', 'wne', 'v', 'm', 'o', 'w10', 'r', 'w', 'wp']}
    namespace = docnsmap if not namespace else namespace
    newelement = etree.Element("{" + nsprefixes[nsprefix] + "}" + tagname,nsmap=namespace)
    if attributes:
        for k in attributes:
            nk = "{%s}"%nsprefixes[attrnsprefix] + k
            full_attr[nk] = attributes[k]
        for a in full_attr:newelement.set(a,full_attr[a])
    if tagtext:newelement.text = tagtext
    return newelement

def addsubelement(parent,tagname,tagtext=None,nsprefix="w",attributes=None, attrnsprefix="w", namespace=None):
    namespace = parent.nsmap if not namespace else namespace
    tag = "{%s}"%nsprefixes[nsprefix]+tagname
    etree.SubElement(parent,tag,nsmap=namespace)
    full_attr = {}
    if attributes:
        for k in attributes:
            nk = "{%s}"%nsprefixes[attrnsprefix] + k
            full_attr[nk] = attributes[k]
        for a in full_attr:parent.getchildren()[-1].set(a,full_attr[a])
    if tagtext: parent.getchildren()[-1].text = tagtext
    return parent

def semicolon_comma_els(symbol):
    a = makeelement("r",nsprefix="w",attributes={"rsidR":"00C14938"},namespace=docnsmap)
    addsubelement(a,"rPr")
    addsubelement(a,"t",tagtext=symbol)
    addsubelement(a.getchildren()[0],"rFonts",attributes={"hint":"eastAsia"})
    return a
  
def samplename_ele(sample):
    a = makeelement("r",nsprefix="w",attributes={"rsidR":"00C14938"},namespace=docnsmap)
    addsubelement(a,"t",tagtext=sample)
    return a
 
def match_str_Element(ele,search,num=0):
    '''Search a ele for a regex, return the matched element, if not matched, return None'''
    result = None
    searchre = re.compile(search)
    for element in ele.iter():
        if element.tag == '{%s}t' % nsprefixes['w']:
            if element.text:
                if searchre.search(element.text):
                    result = element
    if result is not None:
        if range(0,num):
            for i in range(num):
                result = result.getparent()
    return result
    
def matchequal_str_Element(ele,search,num=0):
    '''Search a ele for a regex, return the matched element, if not matched, return None'''
    result = None
    searchre = re.compile(search)
    for element in ele.iter():
        if element.tag == '{%s}t' % nsprefixes['w']:
            if element.text:
                if searchre.search(element.text) and searchre.search(element.text).group() == element.text:
                    result = element
    if result is not None:
        if range(0,num):
            for i in range(num):
                result = result.getparent()
    return result
    
def search(document, search):
    '''Search a document for a regex, return success / fail result'''
    result = False
    searchre = re.compile(search)
    for indec,element in enumerate(document.iter()):
        if element.tag == '{%s}t' % nsprefixes['w']:  # t (text) elements
            if element.text:
                if searchre.search(element.text):
                    print "element index in this document:", index
                    result = True
    return result

def equal_str_Element(ele,search,num=0):
    result = None
    for element in ele.iter():
        if element.tag == '{%s}t' % nsprefixes['w']:
            if element.text:
                if element.text == search:
                    result = element
    if result is not None:
        if range(0,num):
            for i in range(num):
                result = result.getparent()
    return result

def change_sample_num(body_ele,num):
    a = searchElement(body_ele,u"样品数：")
    a.getparent().getnext().getchildren()[0].text = str(num)
    return

def change_project_num(body_ele,name):
    a = matchequalall_str_Element(body_ele,u"项目数据分析报告")
    for i in a:
        i.getparent().getprevious().getchildren()[-1].text = str(name)
    
def coreproperties(project_num, creator=u"生物信息部", author = u"邓勇"):
    """
    Create core properties (common document properties referred to in the
    'Dublin Core' specification). See appproperties() for other stuff.
    """
    import datetime
    core = makeelement('coreProperties',nsprefix="cp",namespace=corensmap)
    addsubelement(core,"title",tagtext=u"%s项目数据分析报告"%project_num,nsprefix="dc")
    addsubelement(core,"creator",tagtext=creator,nsprefix="dc")
    addsubelement(core,"lastModifiedBy",tagtext=author,nsprefix="cp")
    addsubelement(core,"revision",tagtext="1",nsprefix="cp")
    addsubelement(core,"lastPrinted",tagtext=datetime.datetime.today().strftime("%Y-%m-%dT%XZ"),nsprefix="cp")
    addsubelement(core,"created",tagtext=datetime.datetime.today().strftime("%Y-%m-%dT%XZ"),nsprefix="dcterms",attributes=dict(type="dcterms:W3CDTF"),attrnsprefix="xsi")
    addsubelement(core,"modified",tagtext=datetime.datetime.today().strftime("%Y-%m-%dT%XZ"),nsprefix="dcterms",attributes=dict(type="dcterms:W3CDTF"),attrnsprefix="xsi")
    return core
        
def extend(ele,ele_list):
    '''add the element of ele_list as the sibiling element of ele'''
    if isinstance(ele_list,list):
        for e in ele_list[::-1]:ele.addnext(e)
    else:
        try:
            ele.addnext(ele_list)
        except:
            print "not an element"
    return ele
    
def replace(ele, search, replace):
    """
    Replace all occurences of string with a different string, return updated
    document
    """
    searchre = re.compile(search)
    for element in ele.iter():
        if element.tag == '{%s}t' % nsprefixes['w']:  # t (text) elements
            if element.text:
                if searchre.search(element.text):
                    element.text = re.sub(search, replace, element.text)
                    

                    

def modified_info(ele,group_info,vslist,species = "Musm usculus", fa="Mus_musculus.GRCm38.dna.toplevel.fa",gtf="Mus_musculus.GRCm38.84.gtf",egname=None):
    sn = []
    for g in group_info:sn.extend(group_info[g])
    sn = sorted(list(set(sn)))
    egname = sn[0] if not egname else egname
    import re  
    samplenum = match_str_Element(ele,u"样品数：",1)
    samplenum.addnext(samplename_ele(str(len(sn))))
    samplename = match_str_Element(ele,u"样品名称：",1)
    samplenamelist = []
    for s in sn:
        samplenamelist.extend([samplename_ele(s),semicolon_comma_els(u"；")])
    samplenamelist.pop()
    extend(samplename,samplenamelist)
    samplerep = match_str_Element(ele,u"生物学重复样品备注",2)
    if all(len(i) == 1 for i in group_info.values()):
        rep = makeelement("r",attributes=dict(rsidR="002E39F6"))
        addsubelement(rep,"pPr")
        addsubelement(rep.getchildren()[0],"rFonts",attributes=dict(hint="eastAsia"))
        addsubelement(rep,"t",tagtext=u"无")
        samplerep.append(rep)
    else:
        for g in group_info:
            rep = makeelement("p",attributes=dict(rsidR="00C14938",rsidRDefault="00C14938",rsidP="0036148F"))
            addsubelement(rep,"pPr")
            addsubelement(rep.getchildren()[0],"pStyle",attributes=dict(val="12"))
            addsubelement(rep.getchildren()[0],"ind",attributes=dict(firstLine="480"))
            if len(group_info[g]) == 1:
                extend(rep.getchildren()[0],[samplename_ele(g+"="+group_info[g][0])])
            else:
                se = [samplename_ele(g+"="+group_info[g][0])]
                for gs in group_info[g][1:]:
                    se.append(semicolon_comma_els(u"，"))
                    se.append(samplename_ele(gs))
                extend(rep.getchildren()[0],se)
            samplerep.addnext(rep)
    vsgroup = match_str_Element(ele,u"差异基因比较分组信息",1)
    vse = semicolon_comma_els(vslist[0])
    vse.set("rsidR","001E7EB4")
    vsgroup.addnext(vse)
    if len(vslist) > 1:
        x = u"；"
        txt = x.join(vslist[1:])
        vse = semicolon_comma_els(x+txt)
        vse.set("rsidR","00897B53")
        vsgroup.getnext().addnext(vse)
        
    equal_str_Element(ele,"Mus_musculus.GRCm38.dna.toplevel.fa").text = fa
    equal_str_Element(ele,"Mus_musculus.GRCm38.84.gtf").text = gtf
    equal_str_Element(ele,"Musm").text = species.split()[0]
    equal_str_Element(ele,"usculus").text = species.split()[1]
    replace(ele,"examplesamplename",egname)

def group_elemet(ele,module=["clean","align","assemble","expression","cor","deg","deggo","degpathway","function","cluster","snp","novel","asp","corexp","propro"]):
    allmodules = {"clean":u"数据质控","align":u"序列回帖","assemble":u"转录本组装","expression":u"表达量分析","cor":u"样品间相关性分析","deg":u"差异表达基因分析","deggo":u"GO注释和富集","degpathway":u"Pathway注释和富集","function":u"基因功能注释","cluster":u"差异基因表达模式聚类分析","snp":u"cSNP和InDel分析","novel":u"新转录本预测","asp":u"可变剪接分析","corexp":u"差异基因共表达网络分析","propro":u"差异基因蛋白互作网络分析","reference":u"参考文献"}
    l = ["clean","align","assemble","expression","cor","deg","deggo","degpathway","function","cluster","snp","novel","asp","corexp","propro"]
    alleledict=OrderedDict()
    eledict = OrderedDict()
    for m in l:
        m_element = []
        for element in ele.iter():
            if element.tag == '{%s}t' % nsprefixes['w']:
                if element.text:
                    if element.text == allmodules[m]:
                        m_element.append(element)
        for e in m_element:
            if re.search("bookmark",e.getparent().getnext().tag):
                alleledict[m] = e.getparent().getparent()
                break
    for n,v in enumerate(alleledict.values()[:-1]):
        start = ele.index(v)
        end = ele.index(alleledict.values()[n+1])
        alleledict[alleledict.keys()[n]] = ele[start:end]
     
    for c in ele[ele.index(ele[end]):]:
        if equal_str_Element(c,u"参考文献") is not None:
            reference_index = ele.index(c)
            
    alleledict["propro"] = ele[end:reference_index-1]
    alleledict["reference"] = ele[reference_index-1:]
        
    eledict["prefix"] = ele[:ele.index(alleledict["clean"][0])]
    for m in module + ["reference"]:
        if alleledict.has_key(m):eledict[m] = alleledict[m]
    
    return eledict
    
def get_user_name(user=None):
    if user:
        return user
    else:
        userdict={'yongdeng':u"邓  勇","danwu":u"吴  丹","yanpeng":u"彭  岩","haochu":u"褚  浩","zhonghuali":u"李中华","pingmeng":u"孟  萍"}
        user = os.environ["USER"]
        if userdict.has_key(user):
            return userdict[user]
        else:
            print "cannot get username, please modify the output docx file"
    
def get_modules(inputlist):   ### 输入为config.txt中的modules值
    #analysiskey = ["cleandata","align","cufflinks","cuffnorm","cuffdiff","snp","gokegg","coexpression","propro","cluster","asprofile"]
    analysis = {"cleandata":"clean","align":"align","cufflinks":"assemble",
    "cuffnorm":["expression","cor"],"cuffdiff":"deg","snp":"snp","noveltranscripts":"novel",
    "gokegg":["deggo","degpathway","function"],"coexpression":"corexp","propro":"propro","cluster":"cluster","asprofile":"asp"}
    order = ["clean","align","assemble","expression","cor","deg","deggo","degpathway","function","cluster","snp","novel","asp","corexp","propro"]
    modules = []
    for i in inputlist:
        if analysis.has_key(i):
            if isinstance(analysis[i],list):
                modules.extend(analysis[i])
            else:
                modules.append(analysis[i])
    modules.sort(key=order.index)
    return modules

def change_dirname(ele,result_dir):
    dir = next(os.walk(result_dir))[1]
    for d in dir:
        n = re.sub("\d+\.","",d)
        e = matchequalall_str_Element(ele,"\s*\d+\.%s"%n)
        if e:
            for i in e:
                i.text = d
            
def matchequalall_str_Element(ele,search):
    '''Search a ele for a regex, return the matched element, if not matched, return None'''
    result = []
    searchre = re.compile(search)
    for element in ele.iter():
        if element.tag == '{%s}t' % nsprefixes['w']:
            if element.text:
                if searchre.search(element.text) and searchre.search(element.text).group() == element.text:
                    result.append(element)
    return result           


def get_png_url(url):
    page = urllib.urlopen(url)
    html = page.read()
    r = '<img src="(/tmp/.+?\..+?)" '
    s = 'http://www.genome.jp' + re.search(r,html).group(1)
    return s
    
def get_kegg_pic(xlsx,abbr):
    excel = load_workbook(xlsx)
    table1 = excel.get_sheet_by_name("Sheet1")
    rows = table1.max_row
    cols = table1.max_column
    for i in xrange(1,rows+1):
        for j in xrange(1,cols):
            a = str(table1.cell(row=i,column=j).value)
            b = str(table1.cell(row=i,column=j+1).value)
            if a == "KEGG PATHWAY" and b != abbr+"01100":
                url = b + "\n" + str(table1.cell(row=i,column=cols).value)
                return url
                
def get_k01100_pic(xlsx,abbr):
    excel = load_workbook(xlsx)
    table1 = excel.get_sheet_by_name("Sheet1")
    rows = table1.max_row
    cols = table1.max_column
    for i in xrange(1,rows+1):
        for j in xrange(1,cols):
            a = str(table1.cell(row=i,column=j).value)
            b = str(table1.cell(row=i,column=j+1).value)
            if a == "KEGG PATHWAY" and b == abbr+"01100":
                k01100 = str(table1.cell(row=i,column=cols).value)
                return k01100

    
 
def convert_pdf(pdfile,pngfile,resolution=500,type="png"):
    if not os.path.exists(os.path.dirname(pngfile)):os.makedirs(os.path.dirname(pngfile))
    #img = Image(filename=pdfile,resolution=800)
    img = Image(filename=pdfile,background="white",resolution=resolution)
    with img.convert(type) as converted:
        converted.save(filename=pngfile)
    
def group_list(alist,num):
    a = OrderedDict()
    count = 0
    for n,i in enumerate(alist):
        if n % num == 0:
            count += 1
            a.setdefault(str(count),[]).append(i)
        else:
            a.setdefault(str(count),[]).append(i)
    return a.values()

def listdir(path):
    r = []
    for a,b,c in os.walk(path):
        for i in c:
            r.append(os.path.join(a,i))
    return r
 
def renamepic(picpath,picfile):
    n = len(listdir(picpath))
    ext = os.path.splitext(picfile)[-1]
    shutil.copy(picfile, os.path.join(picpath,"image%d%s"%(n+1,ext)))

def renamepic2(infile,outfile):
    ext = os.path.splitext(infile)[-1]
    shutil.copy(infile,outfile+ext)

def get_file_path(dir,suffix):
    for i in listdir(dir):
        if i.endswith(suffix):
            return os.path.abspath(i)

def Schedule(a,b,c):
    '''''
    a:已经下载的数据块
    b:数据块的大小
    c:远程文件的大小
   '''
    per = 100.0 * a * b / c
    if per > 100 :
        per = 100
    print '%.2f%%' % per       
            
def parse_para():
    parser = argparse.ArgumentParser(description="This Script is used to creat the mRNA-Ref project docx file")
    parser.add_argument("-abbr","--abbr",type=str,help="The species abbreviation, like 'hsa'",required = True)
    parser.add_argument("-species","--species",type=str,help="The species name, like 'Homo_sapiens', no space allowed",required = True)
    parser.add_argument("-fa","--fasta",type=str,help="The fasta file basename",required = True)
    parser.add_argument("-gtf","--gtf",type=str,help="The gtf file basename",required = True)
    parser.add_argument("-gs","--group_sample",type=str,help="The groups info, no space can be allowed in some group. the value should be like 'groupA=sampleA1,sampleA2 groupB=sampleB1,sampleB2 ...' ",required = True,nargs="+",metavar="group_info")
    parser.add_argument("-m","--modules",type=str,help='the analysis modules you have done. it should be same as you config.txt file, ["cleandata","align","cufflinks","cuffnorm","cuffdiff","snp","noveltranscripts","gokegg","coexpression","propro","cluster","asprofile"] can be chosen',choices = ["cleandata","align","cufflinks","cuffnorm","cuffdiff","snp","noveltranscripts","gokegg","coexpression","propro","cluster","asprofile"],required = True,nargs="+",metavar="module")
    parser.add_argument("-k","--keep_tmp",action="store_true", default=False,help="Keep the tmp dir, default do not keep")
    parser.add_argument("-o","--output",type=str,help="The output docx file",required = True,metavar="docxfile")
    parser.add_argument("-p","--project_num",type=str,help="project number, like '17-0931'",required = True)
    parser.add_argument("-res","--result_dir",type = str, help="the project results directory, 'pdf_png', 'xlsx_txt' and 'media' dirs will be created under this dir",required = True)
    parser.add_argument("-ppi","--ppi",type=int,help="the resulation of picture in output docxfile, default: 300",default=300)
    parser.add_argument("-vs","--vs",type=str,help="the diff groups, like CasegroupA-VS-ControlgroupB,CasegroupC-VS-ControlgroupD..., '-' symbol must not be in group name", required = True, nargs="+")
    parser.add_argument("-v","--version",action="version",version='%(prog)s 1.0')
    return parser.parse_args()  
            


def clean_table1(elelist,txt):
    tables = [e for e in elelist if e.tag.endswith("tbl")]
    table = tables[-1]
    tabindex = elelist.index(table)
    with open(txt) as fi:
        info = fi.readlines()
        info = [i.split() for i in info]
    tablepre = table[:3]
    sampleraw = table[3]
    raws = min(6,len(info)-1)
    table.extend([sampleraw]*(raws-1))
    l = []
    for i in info[1:raws+1]:l.extend(i)
    for i in table[3:]:
        n = 0
        for e in i.iter():
            if e.text:
                t.text = l[n]
                n+=1
    elelist[tabindex] = table
    
    
def ChangeLastRow(tbl,row=None):
    lastRow = tbl[-1]
    cols = 0
    for i in lastRow.iter():
        if i.text:
            cols += 1
    assert len(row) == cols, "number of columns not equal as the length of given row string list"
    for n,i in enumerate(lastRow.iter()):
        if i.text:
            i.text = row[n]
    return tbl  

def addrows(tbl,row=None):
    lastRow = copy(tbl[-1])
    # bottom = makeelement("bottom",attributes=dict(val="single",sz="4",space="0",color="auto"),namespace=tbl.nsmap)
    # for e in lastRow.iter():
        # if e.tag.endswith("top"):
            # e.addnext(bottom)          ### 将粗线变为细线
    tbl.remove(tbl[-1])
    cols = 0
    new_row = []
    if not isinstance(row[-1],list):     ## add one row:
        new_row.append(row)
    else:
        new_row = copy(row)
    for i in lastRow.iter():
        if i.text and i.text.strip():cols += 1
    assert len(new_row[-1]) == cols, "number of columns not equal as the given row string list"
    for r in new_row:
        ri = iter(r)
        for n,i in enumerate(lastRow.iter()):
            if i.text and i.text.strip():i.text = ri.next()
        tbl.append(lastRow)
        lastRow = copy(tbl[-1])
        
    # for e in tbl[-1].iter():
        # if e.tag.endswith("bottom"):
            # e.getparent().remove(e)            ### 将最后一行细线变为粗线
            
def change_line2(tbl,attrnsprefix="w"):
    tcBorders = makeelement("tcBorders",namespace=tbl.nsmap)
    addsubelement(tcBorders,"top",attributes=dict(val="single",sz="4",space="0",color="auto"))
    #addsubelement(tcBorders,"left",attributes=dict(val="nil"))
    addsubelement(tcBorders,"bottom",attributes=dict(val="single",sz="4",space="0",color="auto"))
    #addsubelement(tcBorders,"right",attributes=dict(val="nil"))
    for r in tbl[2:]:
        for c in r.iter():
            if c.tag == "tcPr":
                if len(c.getchildren()) == 1:  ## 无框线
                    c.append(tcBorders)
                else:
                    tags = [i.tag for i in c[1]]
                    if "bottom" in tags:
                        for d in c[1]:
                            if d.tag == "bottom":
                                c[1].remove(d)
                                c[1].append(makeelement("bottom",namespace=tbl.nsmap,attributes=dict(val="single",sz="4",space="0",color="auto")))
                                #d.attrib["{%s}"%nsprefixes[attrnsprefix]+"sz"] = "4"
                    else:
                        addsubelement(c[1],"bottom",attributes=dict(val="single",sz="4",space="0",color="auto"))
                    if "top" in tags:
                        for d in c[1]:
                            if d.tag == "top":
                                c[1].remove(d)
                                c[1].append(makeelement("top",namespace=tbl.nsmap,attributes=dict(val="single",sz="4",space="0",color="auto")))                            
                            #d.attrib.setdefault("{%s}"%nsprefixes[attrnsprefix]+"sz",4)
                    else:
                        addsubelement(c[1],"top",attributes=dict(val="single",sz="4",space="0",color="auto"))
    for e in tbl[-1].iter():
        if e.tag.endswith("bottom"):e.attrib["{%s}"%nsprefixes[attrnsprefix]+"sz"] = "12"
    for e in tbl[2].iter():
        if e.tag.endswith("top"):e.attrib["{%s}"%nsprefixes[attrnsprefix]+"sz"] = "12"
        
def change_line1(tbl,attrnsprefix="w"):
    tcBorders = makeelement("tcBorders",namespace=tbl.nsmap)
    addsubelement(tcBorders,"top",attributes=dict(val="single",sz="4",space="0",color="auto"))
    addsubelement(tcBorders,"left",attributes=dict(val="nil"))
    addsubelement(tcBorders,"bottom",attributes=dict(val="single",sz="4",space="0",color="auto"))
    addsubelement(tcBorders,"right",attributes=dict(val="nil"))
    for r in tbl[2:]:
        for c in r:
            if c.tag == "tc":
                tcPr = c[0]
                tags = [i.tag for i in tcPr]
                if "tcBorders" in tags:
                    tcPr.remove(tcPr[tags.index("tcBorders")])
                    tcPr.append(tcBorders)
                else:
                    tcPr.append(tcBorders)
    for c in tbl[2]:
        if c.tag == "tc":
            tcPr = c[0]
            tcPr[-1][0].attrib["{%s}"%nsprefixes[attrnsprefix]+"sz"] = "12"

    for c in tbl[-1]:
        if c.tag == "tc":
            tcPr = c[0]
            tcPr[-1][2].attrib["{%s}"%nsprefixes[attrnsprefix]+"sz"] = "12"
            
def change_line2(tbl,attrnsprefix="w"):
    for e in tbl[-1].iter():
        if e.tag.endswith("bottom"):e.attrib["{%s}"%nsprefixes[attrnsprefix]+"sz"] = "12"

def change_line3(tbl,attrnsprefix="w"):
    tcBorders = makeelement("tcBorders",namespace=tbl.nsmap)
    addsubelement(tcBorders,"top",attributes=dict(val="nil"))
    addsubelement(tcBorders,"left",attributes=dict(val="nil"))
    addsubelement(tcBorders,"bottom",attributes=dict(val="single",sz="12",space="0",color="auto"))
    addsubelement(tcBorders,"right",attributes=dict(val="nil"))
    tcBorders1 = makeelement("tcBorders",namespace=tbl.nsmap)
    addsubelement(tcBorders,"top",attributes=dict(val="nil"))
    addsubelement(tcBorders,"left",attributes=dict(val="nil"))
    #addsubelement(tcBorders,"bottom",attributes=dict(val="nil"))
    addsubelement(tcBorders,"right",attributes=dict(val="nil"))
    for r in tbl[4:]:
        for c in r:
            if c.tag == "tc":
                tcPr = c[0]
                tags = [i.tag for i in tcPr]
                if "tcBorders" in tags:
                    tcPr.remove(tcPr[tags.index("tcBorders")])
                tcPr[-1].attrib["{%s}"%nsprefixes[attrnsprefix]+"vAlign"] = "bottom"

def change_line(tbl,attrnsprefix="w"):
    for r in tbl[3:]:
        for c in r:
            if c.tag == "{%s}"%nsprefixes["w"] + "tc":
                tcPr = c[0]
                tags = [i.tag for i in tcPr]
                if "{%s}"%nsprefixes["w"] + "tcBorders" in tags:
                    tcBorders = tcPr[tags.index("{%s}"%nsprefixes["w"]+"tcBorders")]
                    tags2 = [i.tag for i in tcBorders]
                    if "{%s}"%nsprefixes["w"] + "bottom" in tags2:tcBorders.remove(tcBorders[tags2.index("{%s}"%nsprefixes["w"] +"bottom")])
    for c in tbl[-1]:
        if c.tag == "{%s}"%nsprefixes["w"] + "tc":
            tcPr = c[0]
            tags = [i.tag for i in tcPr]
            if "{%s}"%nsprefixes["w"] + "tcBorders" in tags:
                tcPr.remove(tcPr[tags.index("{%s}"%nsprefixes["w"] + "tcBorders")])
    for c in tbl[-1]:
        if c.tag == "{%s}"%nsprefixes["w"] + "tc":
            tc_Borders = makeelement("tcBorders",namespace=c.nsmap)
            addsubelement(tc_Borders,"top",attributes=dict(val="nil"))
            addsubelement(tc_Borders,"left",attributes=dict(val="nil"))
            addsubelement(tc_Borders,"bottom",attributes=dict(val="single",sz="12",space="0",color="auto"))
            addsubelement(tc_Borders,"right",attributes=dict(val="nil"))
            c[0][-1].addnext(tc_Borders)

def addcolnum(tbl,header=None):
    colwidth = copy(tbl[1])
    tbl[1].append(colwidth[-1])
    head = copy(tbl[2][-1])
    for i in head.iter():
        if i.text:i.text = header
    tbl[2].append(head)
    for r in tbl[3:]:
        r.append(copy(r[-1]))
        
def removecolnum(tbl,num):
    for i in range(num):
        tbl[1].remove(tbl[1][-1])
    for i in range(num):
        for r in tbl[2:]:
            r.remove(r[-1])
            
        
        
    

    
# if __name__ == "__main__":
    # main()
    
    
#def main():   #### os.path.abspath(sys.argv[1]) 为结果文件目录
args = parse_para()
name = args.project_num
random_dir = "tmp" + os.popen('echo $RANDOM').read().strip()
if os.path.isdir(random_dir):
    print "%s exists, please used another dirname" %random_dir
    sys.exit(1)
groups = [i.split("=")[0].strip() for i in set(args.group_sample)]
assert len(groups) == len(set(groups)),"Duplicate group name found!"
g_sample = [map(lambda x:x.strip(),i.split("=")[1].strip().split(",")) for i in set(args.group_sample)]
group_info = dict(zip(groups,g_sample))
sn = []
for g in group_info:sn.extend(group_info[g])
sn = sorted(list(set(sn)))

modules_config =  args.modules                    ## config文件中的modules
abbr = args.abbr                               ## config文件中的缩写
vslist = args.vs
d = etree.parse("/lustre/work/yongdeng/RNA_Ref_template/word/document.xml");root = d.getroot();ele = root.getchildren()[0]
modified_info(ele,group_info,vslist,species = args.species.replace("_"," "), fa=os.path.basename(args.fasta) ,gtf=os.path.basename(args.gtf),egname=sn[0])
result = os.path.abspath(args.result_dir)
listresultdir = next(os.walk(result))[1]
for f in listdir(result):
    if f.endswith(".pdf"):
        fp = f.partition(result)
        fn = os.path.join(result,"pdf_png",fp[-1].lstrip("/") + ".png")
        if not os.path.exists(fn):convert_pdf(f,fn,resolution=args.ppi)
    elif f.endswith(".xlsx"):
        fp = f.partition(result)
        fn = os.path.join(result,"xlsx_txt",fp[-1].lstrip("/") + ".txt")
        if not os.path.exists(os.path.dirname(fn)):os.makedirs(os.path.dirname(fn))
        if not os.path.exists(fn):
            try:
                os.system('python /lustre/work/yongdeng/software/python_scripts/tab_csv_xlsx.py -f x2t -i %s -o %s'%(f,fn))
            except:
                continue
module = get_modules(modules_config)   ### 返回类似["clean","align","assemble","expression","cor","deg","deggo","degpathway","function","cluster","snp","novel","asp","corexp","propro"]结果
change_dirname(ele,result)                        ### 修改word中的名字和目录路径对应
wname = get_user_name() if get_user_name() else u"邓  勇"
replace(ele,u"吴  丹",wname)
change_project_num(ele,name)
eg_sn = os.popen("cut -f1 %s"%os.path.join(result,"xlsx_txt","2.cleandata","clean_data.state.xlsx.txt")).read().split()[1:]
eg_sample = sorted(eg_sn[:min(6,len(eg_sn))],key=len,reverse=True)
#for x,y in enumerate(eg_sample):    replace(ele,str(x)+"sample",eg_sample[x])
eledict = group_elemet(ele,module)
media = os.path.join(result,"media")
if not os.path.isdir(media):
    os.makedirs(media)
else:
    shutil.rmtree(media)
    os.makedirs(media)
os.popen("cp /lustre/work/yongdeng/RNA_Ref_template/image*.png %s"%media)



out_ele_list = eledict["prefix"]



if "clean" in module:
    pic1 = get_file_path(result,"rawdata/" + sn[0] + "/" + "R1/per_base_quality.png")   ## 得到图片文件在results目录中的路径
    pic2 = get_file_path(result,"rawdata/" + sn[0] + "/" + "R1/per_base_sequence_content.png")
    pic3 = get_file_path(result,"cleandata/" + sn[0] + "/" + "summary.png")
    renamepic2(pic1,os.path.join(media,"image5"))
    renamepic2(pic2,os.path.join(media,"image6"))
    renamepic2(pic3,os.path.join(media,"image7"))
    
    tbls = [e for e in eledict["clean"] if e.tag.endswith("tbl")]
    rows = []
    qcfile = [i for i in listdir(os.path.join(result,"xlsx_txt")) if os.path.basename(i) == "clean_data.state.xlsx.txt"][0]
    for s in eg_sample:
        s_txt = os.popen("grep -P '^%s\t' %s"%(s,qcfile)).read().strip().split("\t")
        rows.append(s_txt)   
    addrows(tbls[-1],rows)      ## 最后一个表格 
    change_line(tbls[-1])
    
    
    
    out_ele_list.extend(eledict["clean"])
    
if "align" in module:
    pic1 = get_file_path(result,"alignment_statistics/" + sn[0] + "/" + "element_distribution.png")
    pic2 = get_file_path(result,"alignment_statistics/" + sn[0] + "/" + "uniform_distribution.png")
    pic3 = get_file_path(result,"alignment_statistics/" + sn[0] + "/" + "saturation.png")
    renamepic2(pic1,os.path.join(media,"image8"))
    renamepic2(pic2,os.path.join(media,"image9"))
    renamepic2(pic3,os.path.join(media,"image10")) 
    
    tbls = [e for e in eledict["align"] if e.tag.endswith("tbl")]
    rows = []
    qcfile = [i for i in listdir(os.path.join(result,"xlsx_txt")) if os.path.basename(i) == "align.state.xlsx.txt"][0]
    for s in eg_sample:
        s_txt = os.popen("grep -P '^%s\t' %s"%(s,qcfile)).read().strip().split("\t")
        rows.append(s_txt)
    addrows(tbls[-1],rows);change_line(tbls[-1])
    
    out_ele_list.extend(eledict["align"])
    
if "assemble" in module:
    pic = get_file_path(result,"assembly.transcript.length_distribution.png")
    #renamepic(media,pic)
    renamepic2(pic,os.path.join(media,"image11"))
    
    tbls = [e for e in eledict["assemble"] if e.tag.endswith("tbl")]
    rows = []
    qcfile = [i for i in listdir(os.path.join(result,"xlsx_txt")) if os.path.basename(i) == "assembly.state.xlsx.txt"][0]
    for s in eg_sample + ["cuffmerge"]:
        s_txt = os.popen("grep -P '^%s\t' %s"%(s,qcfile)).read().strip().split("\t")
        rows.append(s_txt)
    addrows(tbls[-1],rows);change_line(tbls[-1])
    
    out_ele_list.extend(eledict["assemble"])
    
if "expression" in module:

    tbls = [e for e in eledict["expression"] if e.tag.endswith("tbl")]

    if len(eg_sample) < 6:
        removecolnum(tbls[0],6-len(eg_sample))
        removecolnum(tbls[1],6-len(eg_sample))
    for expfileindex,expfile in enumerate(["gene.expression.xlsx.txt","isoform.expression.xlsx.txt"]):  
        qcfile = [i for i in listdir(os.path.join(result,"xlsx_txt")) if os.path.basename(i) == expfile][0]
        rows = os.popen("head -4 %s"%qcfile).read().strip().split("\n")
        rows1 = [i.split("\t")[:len(eg_sample)+1] for i in rows[1:]]
        addrows(tbls[expfileindex],rows1);change_line(tbls[expfileindex])
        sample_header = rows[0].split("\t")[1:len(eg_sample)+1]
        for x,y in enumerate(sample_header):    replace(tbls[expfileindex],str(x)+"sample",y)
        
    
    qcfile = [i for i in listdir(os.path.join(result,"xlsx_txt")) if os.path.basename(i) == "transid.geneid.refid.pos.xlsx.txt"][0]
    rows1 = os.popen("head -4 %s"%qcfile).read().strip().split("\n")[1:]
    rows1 = [i.split("\t") for i in rows1]
    addrows(tbls[-1],rows1);change_line(tbls[-1])
    
    if "assemble" not in module:
        tbl3_index = eledict["expression"].index(tbls[-1])
        eledict["expression"].pop(tbl3_index+1)
        eledict["expression"].pop(tbl3_index)
        eledict["expression"].pop(tbl3_index-1)
        
        tab1_index = eledict["expression"].index(tbls[0])
        tmp_e = eledict["expression"][tab1_index-3]
        for i in tmp_e.iter():
            if i.text == u"；":
                tmp_ee = i.getparent()
                break
        rm_index = tmp_e.index(tmp_ee)
        for i in tmp_e[rm_index+1:]:
            tmp_e.remove(i)
        tmp_e[-1].remove(tmp_e[-1][-1])
        addsubelement(tmp_e[-1],"t",tagtext=u"。")
           
        

    out_ele_list.extend(eledict["expression"])

if "cor" in module:
    pic = get_file_path(os.path.join(result,"pdf_png"),"sample.correlation.pdf.png")
    renamepic2(pic,os.path.join(media,"image12"))
    
    tbls = [e for e in eledict["cor"] if e.tag.endswith("tbl")]
    if len(eg_sample) < 6:removecolnum(tbls[0],6-len(eg_sample))
    qcfile = [i for i in listdir(os.path.join(result,"xlsx_txt")) if os.path.basename(i) == "sample.correlation.matrix.xlsx.txt"][0]
    rows = os.popen("head -%d %s"%(len(eg_sample)+1,qcfile)).read().strip("\n").split("\n")
    rows1 = [i.split("\t")[:len(eg_sample)+1] for i in rows[1:]]
    rows1 = [[i[0]] + [format(float(j),".6f") for j in i[1:]] for i in rows1]
    addrows(tbls[-1],rows1);change_line(tbls[-1])
    sample_header = rows[0].split("\t")[1:len(eg_sample)+1]
    for x,y in enumerate(sample_header):replace(tbls[-1],str(x)+"sample",y)

    #renamepic(media,pic)
    out_ele_list.extend(eledict["cor"])
if "deg" in module:
    pic1 = get_file_path(os.path.join(result,"pdf_png"),"deg.state.pdf.png")
    pic2 = get_file_path(result,vslist[0] + "/" + vslist[0] + ".plot_scatter.png")
    pic3 = get_file_path(result,vslist[0] + "/" + vslist[0] + ".plot_valcano.png")
    renamepic2(pic1,os.path.join(media,"image13"))
    renamepic2(pic2,os.path.join(media,"image14"))
    renamepic2(pic3,os.path.join(media,"image15"))
    #for p in (pic1,pic2,pic3):renamepic(media,p) 
    out_ele_list.extend(eledict["deg"])
if "deggo" in module:
    pic1 = get_file_path(os.path.join(result,"pdf_png"),"GO/" + vslist[0] + "/" + "deg.go.class.pdf.png")
    pic2 = get_file_path(os.path.join(result,"pdf_png"),"GO/" + vslist[0] + "/" + "enrichment.GO.p_value.top.pdf.png")
    pic3 = sorted([i for i in listdir(os.path.join(result,"pdf_png")) if i.endswith("enrichment.GO.hierarchy.biological_process.pdf.png")],key=lambda x:os.path.getsize(x),reverse=True)[0]
    pic4 = sorted([i for i in listdir(os.path.join(result,"pdf_png")) if i.endswith("enrichment.GO.hierarchy.cellular_component.pdf.png")],key=lambda x:os.path.getsize(x),reverse=True)[0]
    pic5 = sorted([i for i in listdir(os.path.join(result,"pdf_png")) if i.endswith("enrichment.GO.hierarchy.molecular_function.pdf.png")],key=lambda x:os.path.getsize(x),reverse=True)[0]
    # pic1 = get_file_path(os.path.join(result,"pdf_png"),"GO/" + vslist[0] + "/" + "deg.go.class.pdf.png")
    # pic2 = get_file_path(os.path.join(result,"pdf_png"),"GO/" + vslist[0] + "/" + "enrichment.GO.p_value.top.pdf.png")
    # pic3 = get_file_path(os.path.join(result,"pdf_png"),"GO/" + vslist[0] + "/" + "enrichment.GO.hierarchy.biological_process.pdf.png")
    # pic4 = get_file_path(os.path.join(result,"pdf_png"),"GO/" + vslist[0] + "/" + "enrichment.GO.hierarchy.cellular_component.pdf.png")
    # pic5 = get_file_path(os.path.join(result,"pdf_png"),"GO/" + vslist[0] + "/" + "enrichment.GO.hierarchy.molecular_function.pdf.png")
    renamepic2(pic1,os.path.join(media,"image16"))
    renamepic2(pic2,os.path.join(media,"image17"))
    renamepic2(pic3,os.path.join(media,"image18"))
    renamepic2(pic4,os.path.join(media,"image19"))
    renamepic2(pic5,os.path.join(media,"image20"))
    #for p in (pic1,pic2,pic3,pic4,pic5):renamepic(media,p)
    out_ele_list.extend(eledict["deggo"])
if "degpathway" in module:
    pic1 = get_file_path(os.path.join(result,"pdf_png"),"Pathway/" + "KEGG_PATHWAY.second.class.pdf.png")
    renamepic2(pic1,os.path.join(media,"image21"))
    #renamepic(media,pic1)
    pic2 = get_file_path(os.path.join(result,"pdf_png"),"Pathway/" + vslist[0] + "/" + "enrichment.Pathway.p_value.top.pdf.png")
    renamepic2(pic2,os.path.join(media,"image22"))
    #renamepic(media,pic2)
    for f in listdir(result):
        if os.path.basename(f) == "enrichment.Pathway.xlsx":           
            url = get_kegg_pic(f,abbr)
            k01100 = get_k01100_pic(f,abbr)
            print url.split("\n")[-1] + " ------> " + os.path.join(result,"pdf_png",url.split("\n")[0] + ".png")
            print k01100 + " ------> " + os.path.join(result,"pdf_png",abbr+"01100.png")
            if not os.path.isfile(os.path.join(result,"pdf_png",url.split("\n")[0] + ".png")):  ### 后续添加所有kegg pathway链接，下载超时则用下一个
                url_w = get_png_url(url.split("\n")[-1])
                urllib.urlretrieve(url_w,os.path.join(result,"pdf_png",url.split("\n")[0] + ".png"),Schedule)
            if not os.path.isfile(os.path.join(result,"pdf_png",abbr+"01100.png")):
                k01100_w = get_png_url(k01100)
                urllib.urlretrieve(k01100_w,os.path.join(result,"pdf_png",abbr+"01100.png"),Schedule)
            renamepic2(os.path.join(result,"pdf_png",url.split("\n")[0] + ".png"),os.path.join(media,"image23"))
            renamepic2(os.path.join(result,"pdf_png",abbr+"01100.png"),os.path.join(media,"image24"))
            #renamepic(media,os.path.join(result,"pdf_png",url.split("\n")[0] + ".png"))
            #renamepic(media,os.path.join(result,"pdf_png",abbr+"01100.png"))
            break
    out_ele_list.extend(eledict["degpathway"])
if "function" in module:
    out_ele_list.extend(eledict["function"])
if "cluster" in module:
    pic = get_file_path(os.path.join(result,"pdf_png"),"cluster/cluster.samples_heatmap.pdf.png")
    renamepic2(pic,os.path.join(media,"image25"))
    #renamepic(media,pic)
    out_ele_list.extend(eledict["cluster"])
if "snp" in module:

    tbls = [e for e in eledict["snp"] if e.tag.endswith("tbl")]
    qcfile = [i for i in listdir(os.path.join(result,"xlsx_txt")) if os.path.basename(i) == "snp.readcount.xlsx.txt"][0]
    rows = os.popen("head -4 %s"%qcfile).read().strip().split("\n")
    rows1 = [i.split("\t")[:6]+["..."] for i in rows[1:]]
    addrows(tbls[0],rows1);change_line(tbls[0])
    sample_header = rows[0].split("\t")[4].split(".")[0]
    replace(tbls[0],"0sample.Ref",sample_header+".Ref")
    replace(tbls[0],"0sample.Alt",sample_header+".Alt")
    
    
    qcfile = [i for i in listdir(result) if os.path.basename(i) == "snpeff.vcf"][0]
    rows = os.popen("grep -v '^#' %s 2> /dev/null | head -3"%qcfile).read().strip().split("\n")
    rows = [i.split("\t")[:9] for i in rows]
    for i in rows:
        i[7] = ";".join(i[7].split(";")[:2]) + ";..."
        i[8] = "GT:GQ:SDP:DP"
    addrows(tbls[-1],rows);change_line(tbls[-1])

    out_ele_list.extend(eledict["snp"])
if "novel" in module:

    tbls = [e for e in eledict["novel"] if e.tag.endswith("tbl")]
    qcfile = [i for i in listdir(result) if os.path.basename(i) == "novel_transcripts.gtf"][0]
    rows = os.popen("head -3 %s"%qcfile).read().strip().split("\n")
    rows = [i.split("\t") for i in rows]
    addrows(tbls[0],rows);change_line(tbls[0])
    
    qcfile = [i for i in listdir(os.path.join(result,"xlsx_txt")) if os.path.basename(i) == "novel_transcripts.state.xlsx.txt"][0]
    rows=[]
    for s in eg_sample + ["cuffmerge"]:
        s_txt = os.popen("grep -P '^%s\t' %s"%(s,qcfile)).read().strip().split("\t")
        rows.append(s_txt)
    addrows(tbls[-1],rows);change_line(tbls[-1])

    out_ele_list.extend(eledict["novel"])
if "asp" in module:
    pic =  "/lustre/work/yongdeng/RNA_Ref_template/asp.jpeg"                ## 可变剪接图片路径
    renamepic2(pic,os.path.join(media,"image26"))
    
    tbls = [e for e in eledict["asp"] if e.tag.endswith("tbl")]
    qcfile = [i for i in listdir(os.path.join(result,"xlsx_txt")) if os.path.basename(i) == "asprofile.state.xlsx.txt"][0]
    rows=[]
    for s in eg_sample + ["cuffmerge"]:
        s_txt = os.popen("grep -P '^%s\t' %s"%(s,qcfile)).read().strip().split("\t")
        rows.append(s_txt)
    addrows(tbls[-1],rows);change_line(tbls[-1])
    
    #renamepic(media,pic)
    out_ele_list.extend(eledict["asp"])
if "corexp" in module:
    pic = get_file_path(os.path.join(result,"pdf_png"),"coexpression/coexpression.pdf.png")
    if pic:
        renamepic2(pic,os.path.join(media,"image27"))
    else:
        coexp = os.path.join(result,"pdf_png","coexpression.pdf.png")
        print "coexpression picture" + " ------> " + coexp
        if os.path.isfile(coexp):
            renamepic2(coexp,os.path.join(media,"image27"))
            #renamepic(media,coexp)
        else:
            os.system("cp /lustre/work/yongdeng/RNA_Ref_template/image1.png %s"%coexp)
            renamepic2(coexp,os.path.join(media,"image27"))
            #renamepic(media,coexp)
            
    tbls = [e for e in eledict["corexp"] if e.tag.endswith("tbl")]
    qcfile = [i for i in listdir(os.path.join(result,"xlsx_txt")) if os.path.basename(i) == "Co-Expression.xlsx.txt"][0]
    rows = os.popen("head -4 %s"%qcfile).read().strip().split("\n")[1:]
    rows = [i.split("\t") for i in rows]
    
    addrows(tbls[0],rows);change_line(tbls[0])
    
    out_ele_list.extend(eledict["corexp"])
if "propro" in module:
    pic = get_file_path(os.path.join(result,"pdf_png"),vslist[0] + "/" + vslist[0] + ".pdf.png")
    if pic:
        renamepic2(pic,os.path.join(media,"image28"))
    else:
        propic = os.path.join(result,"pdf_png","propro.pdf.png")
        print "pro-pro picture" + " ------> " + propic
        if os.path.isfile(propic):
            renamepic2(propic,os.path.join(media,"image28"))
        else:
            os.system("cp /lustre/work/yongdeng/RNA_Ref_template/image1.png %s"%propic)
            renamepic2(propic,os.path.join(media,"image28"))
            
    tbls = [e for e in eledict["propro"] if e.tag.endswith("tbl")]
    qcfile = [i for i in listdir(os.path.join(result,"xlsx_txt")) if os.path.basename(i) == "deg.pro-pro.xlsx.txt"][0]
    rows = os.popen("head -4 %s"%qcfile).read().strip().split("\n")[1:]
    rows = [i.split("\t")[:-1] for i in rows]
    addrows(tbls[-1],rows);change_line(tbls[-1])
    
    if "corexp" not in module:
        rm_ele = [e for e in eledict["propro"] if match_str_Element(e,u"（说明可见") is not None][0]
        start = rm_ele.index(match_str_Element(rm_ele,u"（说明可见",1))
        end = match_str_Element(rm_ele,u"及附件）",1)
        e_index = rm_ele.index(end)
        for i in rm_ele[start+1:e_index]:
            rm_ele.remove(i)
        end[-1].text =  (end[-1].text).strip(u"及")
            
    out_ele_list.extend(eledict["propro"])

out_ele_list.extend(eledict["reference"])

new_e = makeelement("document",namespace=ele.nsmap) 
new_e.extend(out_ele_list)
core_ele = coreproperties(name, creator=u"生物信息部", author = wname.replace(" ",""))

random_dir = os.path.join(result,random_dir)
os.makedirs(random_dir)
os.system("cp -r /lustre/work/yongdeng/RNA_Ref_template/RNA/* %s"%random_dir) 
os.system("cp -r %s %s"%(media,os.path.join(random_dir,"word")))
with open(os.path.join(random_dir,"docProps","core.xml"),"w") as core: core.write(etree.tostring(core_ele, pretty_print=True,encoding="utf-8",xml_declaration='<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'))
with open(os.path.join(random_dir,"word/document.xml"),"w") as docxml:docxml.write(etree.tostring(new_e, pretty_print=True,encoding="utf-8",xml_declaration='<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'))

os.system("python /lustre/work/yongdeng/software/python_scripts/parsedocx.py %s %s -c"%(random_dir,args.output))

if not args.keep_tmp:
    shutil.rmtree(random_dir)
## etree.tostring(ele, pretty_print=True,encoding="utf-8",xml_declaration='<?xml version="1.0" encoding="UTF-8" standalone="yes"?>')

