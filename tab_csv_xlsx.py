#!/usr/bin/env python
#coding:utf-8

import csv,sys,os
from xlsxwriter.workbook import Workbook
from openpyxl import load_workbook
from collections import OrderedDict
import argparse

def tab2xlsx(tab_file,xlsx_file):
    f1=csv.reader(open(tab_file,"rb"),delimiter="\t")   ### f1为可迭代对象，每迭代一次，返回一个列表，列表为该行分割后形成的列表
    workbook = Workbook(xlsx_file)
    sheet_num = 0
    for r,row in enumerate(f1):
        if r%1048576 == 0:                ### excel一个表中最大的行数为1048576行
            sheet_num+=1
            worksheet = workbook.add_worksheet(name="Sheet%d"%sheet_num)
        for c,col in enumerate(row):
            worksheet.write_string(r%1048576,c,col)   ## 网址超链接将转换为字符， worksheet的write方法会自动识别超链接，若网址字符长度超过255，则会报错，且该内容不会写入到xlsx文件中，write_string将内容强制转换为字符串
    if sheet_num > 1: print "Warning: Too large of the %s file, %d sheets in this file has 1048576 line"%(xlsx_file,sheet_num-1)
    workbook.close()
def tab2csv(tab_file,csv_file):
    f1=csv.reader(open(tab_file,"rb"),delimiter="\t")
    f2=csv.writer(open(csv_file,"wb"))
    for i in f1:
        f2.writerow(i)
        
def csv2tab(csv_file,tab_file):
    f1=csv.reader(open(csv_file,"rb"))
    with open(tab_file,"w") as f2:
        for i in f1:
            f2.write("\t".join(i) + "\n")
            
def csv2xlsx(csv_file,xlsx_file):
    f1=csv.reader(open(csv_file,"rb"))
    workbook = Workbook(xlsx_file)
    worksheet = workbook.add_worksheet()
    for r,row in enumerate(f1):
        for c,col in enumerate(row):
            worksheet.write(r,c,col)

def xlsx2tab_csv(xlsx_file,tab_file,type_out):
    excel = load_workbook(xlsx_file)
    if len(excel.get_sheet_names()) == 1:
        table1 = excel.get_sheet_by_name("Sheet1")    ####只读取第一个表格
        rows = table1.max_row
        cols = table1.max_column
        dict_data = OrderedDict()
        for i in xrange(1,rows+1):
            row_num = "r" + str(i)
            row_list = []
            for j in xrange(1,cols+1):
                if table1.cell(row=i,column=j).value:      #### 当excel表格中数据不为空值时写入
                    row_list.append(str(table1.cell(row=i,column=j).value))
                elif table1.cell(row=i,column=j).value == 0:   ### 当excel表格中数据为0时写入
                    row_list.append("0")
                else:                                      
                    row_list.append("")                       #### 当excel表格中数据为空时(None)，值设为空字符串
            dict_data[row_num] = row_list
    
        if type_out == "t":
            f2 = open(tab_file,"w")
            for k,v in dict_data.iteritems():
                for d in v[:-1]:
                    f2.write(d+"\t")
                f2.write(v[-1] + "\n")
            f2.close()
        elif type_out == "c":
            f2=csv.writer(open(tab_file,"wb"))
            for k,v in dict_data.iteritems():
                f2.writerow(v)
            f2.close()
        nline = int(os.popen("wc -l %s"%tab_file).read().split()[0])
        if nline <= 1:
            print "Warning: no more than one line in output %s file, please check you input file"%tab_file
            os.system("rm -fr %s"%tab_file)
            sys.exit(1)
        return   
        
    for sheet_n in excel.get_sheet_names():  
        table1 = excel.get_sheet_by_name(sheet_n)    #### 读取每一个表格
        rows = table1.max_row
        cols = table1.max_column
        dict_data = OrderedDict()
        for i in xrange(1,rows+1):
            row_num = "r" + str(i)
            row_list = []
            for j in xrange(1,cols+1):
                if table1.cell(row=i,column=j).value:      #### 当excel表格中数据不为空值时写入
                    row_list.append(str(table1.cell(row=i,column=j).value))
                elif table1.cell(row=i,column=j).value == 0:   ### 当excel表格中数据为0时写入
                    row_list.append("0")
                else:                                      
                    row_list.append("")                       #### 当excel表格中数据为空时(None)，值设为空字符串
            dict_data[row_num] = row_list
    
        if type_out == "t":
            f2 = open(os.path.splitext(tab_file)[0] + "_" + sheet_n + os.path.splitext(tab_file)[1],"w")
            for k,v in dict_data.iteritems():
                for d in v[:-1]:
                    f2.write(d+"\t")
                f2.write(v[-1] + "\n")
            f2.close()
        elif type_out == "c":
            f2=csv.writer(open(os.path.splitext(tab_file)[0] + "_" + sheet_n + os.path.splitext(tab_file)[1],"wb"))
            for k,v in dict_data.iteritems():
                f2.writerow(v)
            f2.close()
    nline = int(os.popen("wc -l %s"%tab_file).read().split()[0])
    if nline <= 1:
        print "Warning: no more than one line in output %s file, please check you input file"%tab_file
        os.system("rm -fr %s"%tab_file)
        sys.exit(1)
    return

def parseArgs():
    parser = argparse.ArgumentParser(description="change format among xlsx, csv and tsv")
    parser.add_argument("-f","--change_format",action="store",help="file format of both two, [t==tab or tsv, c==csv, x==xlsx]",choices=["t2c","t2x","c2t","c2x","x2t","x2c"],required = True)
    parser.add_argument("-i","--in_file",action="store",help="the input file",required=True)
    parser.add_argument("-o","--out_file",action="store",help="the output file",required=True)
    args = parser.parse_args()
    return args

def main():
    args = parseArgs()
    in_file = args.in_file
    out_file = args.out_file
    format_str = args.change_format
    format_xlsx = format_str[-1]
    if format_str == "t2c":
        tab2csv(in_file,out_file)
    elif format_str == "t2x":
        tab2xlsx(in_file,out_file)
    elif format_str == "c2t":
        csv2tab(in_file,out_file)
    elif format_str == "c2x":
        csv2xlsx(in_file,out_file)
    #elif format_str == "x2t":
    #    xlsx2tab_csv(in_file,out_file,format_xlsx)
    else:
        xlsx2tab_csv(in_file,out_file,format_xlsx)

if __name__ == '__main__':
    main()
